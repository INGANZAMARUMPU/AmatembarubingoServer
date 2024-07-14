from pprint import pprint
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from django.db.models import Model
from fuzzywuzzy import fuzz
from functools import lru_cache

import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "AmatembarubingoServer.settings")
import django
django.setup()

from api.models import *

wb = load_workbook(sys.argv[1])
sheet = wb.active

corespondances = {
    'Source Aménagée (Isoko ritunganijwe)': SourceAmenagee,
    'Borne fontaine "BF" (IBOMBO RUSANGI )': Ibombo,
    'Borne fontaine "BF" (IBOMBO RUSANGI)': Ibombo,
    'Village collinaire (Ikigwati co ku mutumba)': VillageCollinaire,
    'BRANCHEMENT PRIVE (IMIHANA CANKE INYUBAKWA RUSANGI IFISE AMAZI I WABO )': BranchementPrive,
    'BRANCHEMENT PRIVE (UMUHANA CANKE INYUBAKWA RUZANGI IFISE MAZI IWABO)': BranchementPrive,
    "RESERVOIR (Ikigega c'amazi)": Reservoir,
    'Source Non Aménagée(Isoko ridatunganijwe)': SourceNonAmenagee,
    'Source Non Aménagée (Isoko ridatunganijwe)': SourceNonAmenagee,
    'OUVRAGES AEP': ReseauDAlimentation,
    'CAPTAGE': Captage,
    'SYSTÈME DE POMPAGE': Pompe,
    "PUITS (IRIBA RY'AMAZI RYUBAKIYE)": Puit,
    'FORAGES': Forage,
    'FORAGE': Forage,
    'Village Moderne (Ikigwati ca kijambere)': VillageModerne,
    'BRANCHEMENT PRIVE (UMIHANA CANKE INYUBAKWA RUSANGI IFISE AMAZI I WABO )': BranchementPrive
}

@lru_cache
def predictDBCollumnName(Table:Model, key:str) -> str:
    lower = key.lower()
    if "0. zone d'enqu" in lower : return "II_6_milieu", 100
    if "date" in lower : return "date", 100
    if "ibombo rusangi ryegereye he" in lower : return "IV_1_place", 100
    if "type de branchement prive" in lower : return "V_1_place", 100
    if "prénom" in lower : return "I_1_nom_et_prenom", 100
    if "latitude" in lower : return "II_5_coordonnees", 100
    if "type de branchement prive" in lower : return "V_1_place", 100
    if "borne fontaine proche de" in lower : return "IV_1_place", 100
    
    if not key[0].isalpha() or key[0] != key[0].upper(): return False, 0

    if "aep" in lower : key = "umugende"
    if "identification" in lower : key = "identification"
    if "volume" in lower : key = "volume"
    if "nombre de ménages" in lower : key = "nb_menages"
    if "nombre de menages" in lower : key = "nb_menages"
    if "débit" in lower : key = "debit"
    if "debit" in lower : key = "debit"
    if "ibara" in lower : key = "coloration"
    if "tarissement" in lower : key = "tarissement"
    if "agacimbiri" in lower : key = "sous_colline"
    if "ivyihwejwe" in lower : key = "observations"
    ratio = 0
    prediction = ""
    for column in [x.name for x in Table._meta.get_fields(include_hidden=True)]:
        if not prediction:
            prediction = column
            ratio = fuzz.ratio(key, column)
            continue
        new_ratio = fuzz.ratio(key, column)
        if new_ratio > ratio:
            ratio = new_ratio
            prediction = column
    return prediction, ratio

def forceBoolean(valeur:str):
    try:
        lower = valeur.lower()
    except Exception:
        return valeur
    if "(ntirikora)" in lower or "(oya)" in lower or "nom" == lower:
        return False
    if "irakora)" in lower or "(ego)" in lower or "oui" == lower:
        return True
    return valeur

for i, ligne in tqdm(enumerate(sheet.iter_rows(min_row=2, values_only=True))):
    i += 1
    data = {}
    confidences = {}
    table_name = sheet[f"A{i+1}"].value
    table:Model = corespondances[table_name]
    for j, col in enumerate(ligne):
        letter = get_column_letter(j+1)
        if letter == "A": continue
        valeur = sheet[f"{letter}{i+1}"].value
        if valeur == "-": continue
        valeur = forceBoolean(valeur)
        column_name = sheet[f"{letter}1"].value
        db_column_name, confidence = predictDBCollumnName(table, column_name)
        if db_column_name:
            if not data.get(db_column_name):
                data[db_column_name] = valeur
                confidences[db_column_name] = confidence
            else:
                if confidences[db_column_name] < confidence:
                    data[db_column_name] = valeur
                    confidences[db_column_name] = confidence
    copie = dict(data)
    try:
        to_delete = ["date"]
        for key in data.keys():
            if data[key] == None: to_delete.append(key)
        for x in to_delete:
            del data[x]
        table.objects.create(**data)
    except Exception as e:
        pprint(copie)
        print(e)

