import openpyxl
from tqdm import tqdm

conditions = {
    "BF": {
        "rural": {"index_debit": [], "index_menages": []},
        "urbain": {"index_debit": [], "index_menages": []},
    },
    "PRIVE": {
        "rural": {"index_debit": [], "index_menages": []},
        "urbain": {"index_debit": [], "index_menages": []},
    },
    "CAPTAGE": {
        "rural": {"index_debit": [], "index_menages": []},
        "urbain": {"index_debit": [], "index_menages": []},
    },
    "FORAGES": {
        "rural": {"index_debit": [], "index_menages": []},
        "urbain": {"index_debit": [], "index_menages": []},
    },
    "AEP": {
        "rural": {"index_debit": [], "index_menages": []},
        "urbain": {"index_debit": [], "index_menages": []},
    },
    "PUITS": {
        "rural": {"index_debit": [], "index_menages": []},
        "urbain": {"index_debit": [], "index_menages": []},
    },
    "RESERVOIR": {
        "rural": {"index_debit": [], "index_menages": []},
        "urbain": {"index_debit": [], "index_menages": []},
    },
    "ritunganijwe": {
        "rural": {"index_debit": [], "index_menages": []},
        "urbain": {"index_debit": [], "index_menages": []},
    },
    "ridatunganijwe": {
        "rural": {"index_debit": [], "index_menages": []},
        "urbain": {"index_debit": [], "index_menages": []},
    },
    "POMPAGE": {
        "rural": {"index_debit": [], "index_menages": []},
        "urbain": {"index_debit": [], "index_menages": []},
    },
}

def merge_excel_files(file):
    CORRECTS = []
    WRONGS = []
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    titles = []
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        for item in row:
            titles.append(item)

    for row in tqdm(sheet.iter_rows(min_row=2, values_only=True)):
        for key in conditions.keys():
            if key in row["Q"]:
                value = conditions[key]
                if "rural" in row["C"]:
                    wrong = False
                    for cle in value["rural"].keys:
                        valeur = value["rural"][cle]
                        if len(valeur) == 0: continue
                        if row[cle] < valeur[0] or row[cle] > valeur[1]:
                            wrong = True
                            continue
                    if wrong:
                        WRONGS.append(row)
                    else:
                        CORRECTS.append(row)
                else:
                    wrong = False
                    for cle in value["urbain"].keys:
                        valeur = value["urbain"][cle]
                        if len(valeur) == 0: continue
                        if row[cle] < valeur[0] or row[cle] > valeur[1]:
                            wrong = True
                            continue
                    if wrong:
                        WRONGS.append(row)
                    else:
                        CORRECTS.append(row)


    correct_wb = openpyxl.Workbook()
    correct_sheet = correct_wb.active
    correct_sheet.append(titles)
    for data in tqdm(CORRECTS):
        correct_sheet.append(data)
    correct_wb.save("correct_file.xlsx")

    wrong_wb = openpyxl.Workbook()
    wrong_sheet = wrong_wb.active
    wrong_sheet.append(titles)
    for data in tqdm(WRONGS):
        wrong_sheet.append(data)
    wrong_wb.save("wrong_file.xlsx")

merge_excel_files("merged_file.xlsx")
