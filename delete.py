from pprint import pprint
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm
from django.db.models import Model
from functools import lru_cache
from pathlib import Path

import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "AmatembarubingoServer.settings")
import django
django.setup()

from api.models import *

corespondances = {
    'Source Aménagée (Isoko ritunganijwe)': SourceAmenagee,
    'Borne fontaine "BF" (IBOMBO RUSANGI)': Ibombo,
    'Village collinaire (Ikigwati co ku mutumba)': VillageCollinaire,
    'BRANCHEMENT PRIVE (IMIHANA CANKE INYUBAKWA RUSANGI IFISE AMAZI I WABO )': BranchementPrive,
    "RESERVOIR (Ikigega c'amazi)": Reservoir,
    'Source Non Aménagée (Isoko ridatunganijwe)': SourceNonAmenagee,
    'OUVRAGES AEP': ReseauDAlimentation,
    'CAPTAGE': Captage,
    'SYSTÈME DE POMPAGE': Pompe,
    "PUITS (IRIBA RY'AMAZI RYUBAKIYE)": Puit,
    'FORAGE': Forage,
    'Village Moderne (Ikigwati ca kijambere)': VillageModerne,
    'BRANCHEMENT PRIVE (UMUHANA CANKE INYUBAKWA RUZANGI IFISE MAZI IWABO)': BranchementPrive
}
def process(sheet):
    for i, ligne in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        i += 1
        table_name = sheet[f"C{i+1}"].value
        id = sheet[f"E{i+1}"].value
        try:
            table:Model = corespondances[table_name]
            objects = table.objects.filter(id = id)
        except Exception as e:
            pass
        objects.delete()

for file in os.listdir("files"):
    wb = load_workbook(f"files/{file}")
    sheet = wb.active
    process(sheet)