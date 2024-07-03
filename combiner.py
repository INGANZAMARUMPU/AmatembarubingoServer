from pathlib import *
import openpyxl
from tqdm import tqdm

def extract_titles(files):
    merged_data = set()
    titles = []
    
    # Parcourir tous les fichiers
    for file in files:
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        print(file)
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for item in row:
                if item not in merged_data:
                    merged_data.add(item)
                    titles.append(item)

    return(titles)

def merge_excel_files(files, columns:list):
    DB = []
    for file in files:
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        titles = []
        for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for item in row:
                titles.append(item)

        for row in tqdm(sheet.iter_rows(min_row=2, values_only=True)):
            DATA = ["-" for _ in columns]
            for i, item in enumerate(row):
                title = titles[i]
                db_index = columns.index(title)
                if item or item == 0: DATA[db_index] = item 
            DB.append(DATA)
    
    # Créer un nouveau classeur et feuille de calcul pour les données fusionnées
    merged_wb = openpyxl.Workbook()
    merged_sheet = merged_wb.active
    
    merged_sheet.append(columns)
    
    for data in tqdm(DB):
        merged_sheet.append(data)
    
    # Enregistrer le fichier fusionné
    merged_wb.save("merged_file.xlsx")

files = [x.name for x in Path("brouillons").iterdir()]

titles = extract_titles(files)
merge_excel_files(files, titles)
